import openpyxl as op
from openpyxl.styles import Font
from os.path import exists


def common_keywords(cal_val):
    data.cell(row=1, column=cal_val + 3).font = Font(bold=True)
    if cal_val == 3:
        data.cell(row=1, column=cal_val + 3).value = "Component based on problem statement"
    elif cal_val == 4:
        data.cell(row=1, column=cal_val + 3).value = "Component based on Original statement"
    for i in range(2, total_data_rows + 1):
        problem_description = data.cell(row=i, column=cal_val).value.lower()
        temp = []
        for keyword in keywords:
            if keyword in problem_description:
                temp.append(keyword)
        temp = ",".join(temp)
        data.cell(row=i, column=cal_val + 3).value = temp
    content.save(file)
    global flag
    flag = flag + 1
    if flag == 2:
        data.cell(row=1, column=8).value = "Common from F and G"
        for i in range(2, total_data_rows + 1):
            problem = data.cell(row=i, column=6).value.split(",")
            problem = set(problem)
            original = data.cell(row=i, column=7).value.split(",")
            original = set(original)
            if problem & original:
                temp = problem & original
                temp = list(temp)
                temp = ','.join(str(z) for z in temp)
                data.cell(row=i, column=8).value = temp
        content.save(file)


def file_exists(file):
    if not exists(file):
        print('\nFile not found\nCheck for file path or file name.\n')
        exit(1)
    else:
        return True


file = input("\nEnter excel filename: ")
if 'xlsx' in file:
    file_exists(file)
else:
    file = file + '.xlsx'
    file_exists(file)


data_sheet_name = input("\nEnter sheet name(Case Volume): ")
component_sheet_name = input("\nEnter sheet name(Keywords): ")

print("\nIn progress...\n")
content = op.load_workbook(file)

data = content[data_sheet_name]
components = content[component_sheet_name]

total_data_rows = data.max_row
total_keyword_rows = components.max_row

keywords = [components.cell(row=i, column=1).value.lower() for i in range(2, total_keyword_rows +
                                                                          1)]
flag = 0

for i in (3, 4):
    common_keywords(i)
print("Done")
exit(0)


